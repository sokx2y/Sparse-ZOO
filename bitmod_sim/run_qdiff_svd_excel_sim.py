#!/usr/bin/env python3
"""
Run native bitmod-sim/BitFusion single-linear hardware modeling for QdiffSVDLinear
rows described by an Excel/CSV table.

The script reads the workbook layout used by NEWHardwareResult.xlsx:
one parameter row per model/layer followed by operation rows:
  w*x, w*dx, svd_out, svd_x, svd_dx

It writes:
  - summary CSV: one row per Excel operation row with aggregated metrics
  - detail CSV: one row per internal GEMM that was summed
  - optional filled XLSX copy when openpyxl is available and the input is XLSX
"""

from __future__ import annotations

import argparse
import ast
import copy
import csv
import math
import re
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


RESULT_COLUMNS = [
    "compute_latency(cycles)",
    "dram_latency",
    "total latency",
    "compute energy(uJ)",
    "sram rd/wr energy",
    "dram energy",
    "on-chip energy",
    "total energy",
]

DRAM_PROFILE_DATA_RATE_MTPS = {
    "ddr4-3200": 3200.0,
    "ddr5-4800": 4800.0,
    "ddr5-5600": 5600.0,
    "ddr5-6400": 6400.0,
}


def effective_dram_bw_bits_per_cycle(args: argparse.Namespace) -> float:
    if args.dram_data_rate_mtps is None and args.dram_profile == "legacy":
        return 256.0
    data_rate_mtps = (
        args.dram_data_rate_mtps
        if args.dram_data_rate_mtps is not None
        else DRAM_PROFILE_DATA_RATE_MTPS[args.dram_profile]
    )
    return (
        float(data_rate_mtps)
        * float(args.dram_bus_width_bits)
        * int(args.dram_channels)
        / float(args.accelerator_freq_mhz)
    )

PARAM_COLUMNS = [
    "model",
    "layer",
    "linear",
    "BS",
    "CXT_LEN",
    "r",
    "in_feature",
    "out_feature",
    "svd_rank",
    "w_prec",
    "i_prec",
    "o_prec",
]


@dataclass(frozen=True)
class GemmSpec:
    name: str
    batch_size: int
    cxt_len: int
    in_features: int
    out_features: int
    w_prec: float
    i_prec: float
    output_prec: float


def _norm_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _nonempty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def _to_int(value: Any, name: str) -> int:
    if value is None or value == "":
        raise ValueError(f"Missing required integer field: {name}")
    return int(float(value))


def _to_float(value: Any, name: str) -> float:
    if value is None or value == "":
        raise ValueError(f"Missing required precision field: {name}")
    return float(value)


def _read_csv(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = [_norm_header(h) for h in (reader.fieldnames or [])]
        rows = []
        for idx, row in enumerate(reader, start=2):
            normalized = {_norm_header(k): v for k, v in row.items()}
            normalized["__rownum__"] = idx
            rows.append(normalized)
    return headers, rows


def _read_xlsx(path: Path, sheet_name: Optional[str]) -> Tuple[List[str], List[Dict[str, Any]], Any, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "Reading .xlsx requires openpyxl. Install it in the Linux env with:\n"
            "  pip install openpyxl\n"
            "or export the sheet as CSV and pass that CSV to this script."
        ) from exc

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = [_norm_header(cell.value) for cell in sheet[1]]
    rows: List[Dict[str, Any]] = []
    for row_idx in range(2, sheet.max_row + 1):
        record = {headers[col_idx - 1]: sheet.cell(row_idx, col_idx).value for col_idx in range(1, len(headers) + 1)}
        record["__rownum__"] = row_idx
        rows.append(record)
    return headers, rows, workbook, sheet


def read_table(path: Path, sheet_name: Optional[str]) -> Tuple[List[str], List[Dict[str, Any]], Any, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        headers, rows = _read_csv(path)
        return headers, rows, None, None
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path, sheet_name)
    raise SystemExit(f"Unsupported input format: {path.suffix}. Use .xlsx, .xlsm, or .csv.")


_SIMULATOR_CLASS = None


def get_cached_simulator_class():
    """Import bitmod-sim lazily so --help works before torch is installed/active."""

    global _SIMULATOR_CLASS
    if _SIMULATOR_CLASS is not None:
        return _SIMULATOR_CLASS

    try:
        from mem.mem_instance import MemoryInstance
        from single_linear_sim_bitFusion import SingleLinearSimulatorBB
    except ModuleNotFoundError as exc:
        missing = exc.name or str(exc)
        raise SystemExit(
            f"Missing Python dependency while loading bitmod-sim: {missing}\n"
            "Activate the Linux conda environment that contains torch, or install the missing package."
        ) from exc

    class CachedSingleLinearSimulatorBB(SingleLinearSimulatorBB):
        """Reuse CACTI-derived memory instances for identical SRAM configs."""

        _mem_cache: Dict[Tuple[float, float, float, float, int, int, int, int, float], Tuple[Any, Any, Any]] = {}

        def _init_mem(self):
            if self.is_bit_serial:
                w_bandwidth = self.pe_dp_size * math.ceil(float(self.w_prec) / 4) * 4 * self.pe_array_dim["h"] / 2
                i_bandwidth = self.pe_dp_size * float(self.i_prec) * self.pe_array_dim["w"] / 2
            else:
                w_bandwidth = self.pe_dp_size * math.ceil(float(self.w_prec) / 4) * 4 * self.pe_array_dim["h"]
                i_bandwidth = self.pe_dp_size * float(self.i_prec) * self.pe_array_dim["w"]

            key = (
                float(self.w_prec),
                float(self.i_prec),
                float(self.output_prec),
                float(self.pe_dp_size),
                int(self.pe_array_dim["h"]),
                int(self.pe_array_dim["w"]),
                int(w_bandwidth),
                int(i_bandwidth),
                0.028,
            )
            cached = self._mem_cache.get(key)
            if cached is not None:
                self.w_sram, self.i_sram, self.dram = copy.deepcopy(cached)
                return

            w_sram_config = {
                "technology": 0.028,
                "mem_type": "ram",
                "size": 512 * 1024 * 8,
                "bank_count": 8,
                "rw_bw": w_bandwidth,
                "r_port": 1,
                "w_port": 1,
                "rw_port": 0,
            }
            self.w_sram = MemoryInstance(
                w_sram_config,
                r_cost=0,
                w_cost=0,
                latency=1,
                min_r_granularity=None,
                min_w_granularity=64,
                get_cost_from_cacti=True,
            )

            i_sram_config = {
                "technology": 0.028,
                "mem_type": "ram",
                "size": 512 * 1024 * 8,
                "bank_count": 8,
                "rw_bw": i_bandwidth,
                "r_port": 1,
                "w_port": 1,
                "rw_port": 0,
            }
            self.i_sram = MemoryInstance(
                i_sram_config,
                r_cost=0,
                w_cost=0,
                latency=1,
                min_r_granularity=64,
                min_w_granularity=64,
                get_cost_from_cacti=True,
            )

            dram_rw_bw = 128
            dram_config = {
                "technology": 0.028,
                "mem_type": "dram",
                "size": 1e9 * 8,
                "bank_count": 1,
                "rw_bw": dram_rw_bw,
                "r_port": 0,
                "w_port": 0,
                "rw_port": 1,
            }
            wr_cost = dram_rw_bw / 64 * 1200
            self.dram = MemoryInstance(
                dram_config,
                r_cost=wr_cost,
                w_cost=wr_cost,
                latency=1,
                min_r_granularity=dram_rw_bw,
                min_w_granularity=dram_rw_bw,
                get_cost_from_cacti=False,
            )
            self._mem_cache[key] = copy.deepcopy((self.w_sram, self.i_sram, self.dram))

    _SIMULATOR_CLASS = CachedSingleLinearSimulatorBB
    return _SIMULATOR_CLASS


def simulate_gemm(spec: GemmSpec, args: argparse.Namespace) -> Dict[str, Any]:
    if args.backend == "script":
        return simulate_gemm_with_script(spec, args)

    simulator_class = get_cached_simulator_class()
    x_shape = (spec.batch_size, spec.cxt_len, spec.in_features)
    w_shape = (spec.out_features, spec.in_features)
    sim = simulator_class(
        x=x_shape,
        w=w_shape,
        i_prec=spec.i_prec,
        kv_prec=args.kv_prec,
        w_prec=spec.w_prec,
        output_prec=spec.output_prec,
        batch_size=spec.batch_size,
        is_bit_serial=args.is_bit_serial,
        pe_dp_size=args.pe_dp_size,
        pe_energy=args.pe_energy,
        pe_area=args.pe_area,
        pe_array_dim=(args.pe_array_h, args.pe_array_w),
        init_mem=True,
        cxt_len=spec.cxt_len,
        is_generation=False,
        layer_name="single_linear",
        base_activation_prec=args.base_activation_prec,
        base_weight_prec=args.base_weight_prec,
        dram_effective_bw_bits_per_cycle=effective_dram_bw_bits_per_cycle(args),
        dram_energy_pj_per_bit=args.dram_energy_pj_per_bit,
    )
    return sim.simulate()


def _parse_required(pattern: str, text: str, name: str) -> str:
    match = re.search(pattern, text)
    if not match:
        raise RuntimeError(f"Could not parse {name} from simulator output:\n{text}")
    return match.group(1)


def simulate_gemm_with_script(spec: GemmSpec, args: argparse.Namespace) -> Dict[str, Any]:
    script_path = args.sim_script.resolve()
    cmd = [
        sys.executable,
        str(script_path),
        "--batch_size",
        str(spec.batch_size),
        "--cxt_len",
        str(spec.cxt_len),
        "--in_features",
        str(spec.in_features),
        "--out_features",
        str(spec.out_features),
        "--i_prec",
        str(spec.i_prec),
        "--kv_prec",
        str(args.kv_prec),
        "--w_prec",
        str(spec.w_prec),
        "--output_prec",
        str(spec.output_prec),
        "--pe_dp_size",
        str(args.pe_dp_size),
        "--pe_energy",
        str(args.pe_energy),
        "--pe_area",
        str(args.pe_area),
        "--pe_array_h",
        str(args.pe_array_h),
        "--pe_array_w",
        str(args.pe_array_w),
        "--base_activation_prec",
        str(args.base_activation_prec),
        "--base_weight_prec",
        str(args.base_weight_prec),
        "--dram_profile",
        str(args.dram_profile),
        "--dram_bus_width_bits",
        str(args.dram_bus_width_bits),
        "--dram_channels",
        str(args.dram_channels),
        "--accelerator_freq_mhz",
        str(args.accelerator_freq_mhz),
        "--dram_energy_pj_per_bit",
        str(args.dram_energy_pj_per_bit),
    ]
    if args.dram_data_rate_mtps is not None:
        cmd.extend(["--dram_data_rate_mtps", str(args.dram_data_rate_mtps)])
    if args.is_bit_serial:
        cmd.append("--is_bit_serial")

    completed = subprocess.run(
        cmd,
        cwd=script_path.parent,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(
            f"Simulator failed for {spec.name} with return code {completed.returncode}.\n"
            f"Command: {' '.join(cmd)}\n"
            f"Output:\n{completed.stdout}"
        )

    out = completed.stdout
    gemm_match = re.search(r"gemm\(M,N,K\)=\((\d+),\s*(\d+),\s*(\d+)\)", out)
    if not gemm_match:
        raise RuntimeError(f"Could not parse GEMM shape from simulator output:\n{out}")
    memory_match = re.search(r"memory bytes:\s+(\{[^\n]+\})", out)
    refetch_match = re.search(r"num mem refetch:\s+(\{[^\n]+\})", out)
    if not memory_match:
        raise RuntimeError(f"Could not parse memory bytes from simulator output:\n{out}")
    if not refetch_match:
        raise RuntimeError(f"Could not parse num mem refetch from simulator output:\n{out}")

    def parse_float(label: str) -> float:
        return float(_parse_required(rf"{re.escape(label)}:\s+([-+0-9.eE]+)", out, label))

    compute_cycles = parse_float("compute latency")
    dram_cycles = parse_float("dram latency")
    total_cycles = parse_float("total latency")
    compute_energy_uJ = parse_float("compute energy")
    sram_rd_uJ = parse_float("sram rd energy")
    sram_wr_uJ = parse_float("sram wr energy")
    dram_energy_uJ = parse_float("dram energy")
    onchip_uJ = parse_float("on-chip energy")
    total_uJ = parse_float("total energy")

    return {
        "gemm_shape": {
            "m": int(gemm_match.group(1)),
            "n": int(gemm_match.group(2)),
            "k": int(gemm_match.group(3)),
        },
        "cycle": {
            "compute": compute_cycles,
            "dram": dram_cycles,
            "total": total_cycles,
        },
        "energy_pj": {
            "compute": compute_energy_uJ * 1e6,
            "sram_rd": sram_rd_uJ * 1e6,
            "sram_wr": sram_wr_uJ * 1e6,
            "dram": dram_energy_uJ * 1e6,
            "onchip": onchip_uJ * 1e6,
            "total": total_uJ * 1e6,
        },
        "memory_bytes": ast.literal_eval(memory_match.group(1)),
        "num_mem_refetch": ast.literal_eval(refetch_match.group(1)),
    }


def build_gemm_specs(op: str, ctx: Dict[str, Any]) -> List[GemmSpec]:
    bs = _to_int(ctx["BS"], "BS")
    cxt_len = _to_int(ctx["CXT_LEN"], "CXT_LEN")
    r = _to_int(ctx["r"], "r")
    in_feature = _to_int(ctx["in_feature"], "in_feature")
    out_feature = _to_int(ctx["out_feature"], "out_feature")
    svd_rank = _to_int(ctx["svd_rank"], "svd_rank")
    w_prec = _to_float(ctx["w_prec"], "w_prec")
    i_prec = _to_float(ctx["i_prec"], "i_prec")
    output_prec = _to_float(ctx["o_prec"], "o_prec")

    def main(name: str, in_f: int, out_f: int) -> GemmSpec:
        return GemmSpec(name, bs, cxt_len, in_f, out_f, w_prec, i_prec, output_prec)

    def factor(name: str, rows: int, cols: int) -> GemmSpec:
        # Low-rank provider forms a parameter delta as left @ right.T:
        # [rows, r] @ [r, cols] -> [rows, cols].
        return GemmSpec(name, 1, rows, r, cols, w_prec, i_prec, output_prec)

    op_key = op.strip().lower()
    if op_key in {"w*x", "wx"}:
        return [main("w_x", in_feature, out_feature)]
    if op_key in {"w*dx", "wdx"}:
        return [main("w_dx", in_feature, out_feature)]
    if op_key == "svd_out":
        return [
            main("svd_out_v_x", in_feature, svd_rank),
            main("svd_out_u_tmp", svd_rank, out_feature),
        ]
    if op_key == "svd_x":
        return [
            factor("svd_x_make_du", out_feature, svd_rank),
            factor("svd_x_make_dv", in_feature, svd_rank),
            main("svd_x_v_plus_dv_x", in_feature, svd_rank),
            main("svd_x_u_plus_du_tmp", svd_rank, out_feature),
        ]
    if op_key == "svd_dx":
        return [
            factor("svd_dx_make_du", out_feature, svd_rank),
            factor("svd_dx_make_dv", in_feature, svd_rank),
            main("svd_dx_v_plus_dv_dx", in_feature, svd_rank),
            main("svd_dx_u_plus_du_tmp", svd_rank, out_feature),
        ]
    raise ValueError(f"Unsupported linear operation: {op}")


def empty_metrics() -> Dict[str, float]:
    return {
        "compute_latency(cycles)": 0.0,
        "dram_latency": 0.0,
        "total latency": 0.0,
        "compute energy(uJ)": 0.0,
        "sram rd/wr energy": 0.0,
        "dram energy": 0.0,
        "on-chip energy": 0.0,
        "total energy": 0.0,
    }


def metrics_from_result(result: Dict[str, Any]) -> Dict[str, float]:
    energy_uJ = {name: value / 1e6 for name, value in result["energy_pj"].items()}
    return {
        "compute_latency(cycles)": float(result["cycle"]["compute"]),
        "dram_latency": float(result["cycle"]["dram"]),
        "total latency": float(result["cycle"]["total"]),
        "compute energy(uJ)": energy_uJ["compute"],
        "sram rd/wr energy": energy_uJ["sram_rd"] + energy_uJ["sram_wr"],
        "dram energy": energy_uJ["dram"],
        "on-chip energy": energy_uJ["onchip"],
        "total energy": energy_uJ["total"],
    }


def add_metrics(dst: Dict[str, float], src: Dict[str, float]) -> None:
    for key in RESULT_COLUMNS:
        dst[key] += src[key]


def _result_component_dram_costs(
    result: Dict[str, Any],
    component: str,
    effective_bw_bits_per_cycle: float,
    energy_pj_per_bit: float,
) -> Tuple[float, float]:
    num_bytes = float(result["memory_bytes"].get(component, 0.0))
    refetch = 1.0
    if component != "output":
        refetch = float(result["num_mem_refetch"].get(component, 1.0))
    return (
        dram_cycles_for_bytes(num_bytes, effective_bw_bits_per_cycle, refetch),
        dram_energy_uJ_for_bytes(num_bytes, energy_pj_per_bit, refetch),
    )


def apply_dram_component_adjustment(
    result: Dict[str, Any],
    metrics: Dict[str, float],
    effective_bw_bits_per_cycle: float,
    energy_pj_per_bit: float,
    remove_components: Sequence[str],
) -> Dict[str, float]:
    adjusted = dict(metrics)
    removed_cycles = 0.0
    removed_energy = 0.0

    for component in remove_components:
        cycles, energy = _result_component_dram_costs(
            result, component, effective_bw_bits_per_cycle, energy_pj_per_bit
        )
        removed_cycles += cycles
        removed_energy += energy

    old_dram = adjusted["dram_latency"]
    old_total = adjusted["total latency"]
    new_dram = max(0.0, old_dram - removed_cycles)
    new_total = max(adjusted["compute_latency(cycles)"], new_dram)

    adjusted["dram_latency"] = new_dram
    adjusted["total latency"] = new_total
    adjusted["dram energy"] = max(0.0, adjusted["dram energy"] - removed_energy)
    adjusted["total energy"] = adjusted["on-chip energy"] + adjusted["dram energy"]
    adjusted["_dram_latency_removed"] = old_dram - new_dram
    adjusted["_total_latency_removed"] = old_total - new_total
    adjusted["_dram_energy_removed"] = removed_energy
    adjusted["_removed_components"] = "+".join(remove_components)
    return adjusted


def dram_cycles_for_bytes(
    num_bytes: float, effective_bw_bits_per_cycle: float, refetch: float = 1.0
) -> float:
    return num_bytes * 8 / effective_bw_bits_per_cycle * refetch


def dram_energy_uJ_for_bytes(
    num_bytes: float, energy_pj_per_bit: float, refetch: float = 1.0
) -> float:
    return num_bytes * 8 * energy_pj_per_bit * refetch / 1e6


def apply_uv_tmp_sram_adjustment(
    spec: GemmSpec,
    result: Dict[str, Any],
    metrics: Dict[str, float],
    effective_bw_bits_per_cycle: float,
    energy_pj_per_bit: float,
    remove_output_tmp: bool = False,
    remove_input_tmp: bool = False,
) -> Dict[str, float]:
    remove_components = []
    if remove_output_tmp:
        remove_components.append("output")
    if remove_input_tmp:
        remove_components.append("input")
    return apply_dram_component_adjustment(
        result,
        metrics,
        effective_bw_bits_per_cycle,
        energy_pj_per_bit,
        remove_components,
    )


def uv_tmp_sram_roles(specs: Sequence[GemmSpec]) -> Dict[int, Tuple[bool, bool]]:
    roles: Dict[int, Tuple[bool, bool]] = {}
    for idx in range(len(specs) - 1):
        first = specs[idx].name
        second = specs[idx + 1].name
        if (
            ("_v_" in first or "_v_plus_dv_" in first)
            and ("_u_" in second or "_u_plus_du_" in second)
            and specs[idx].out_features == specs[idx + 1].in_features
            and specs[idx].batch_size == specs[idx + 1].batch_size
            and specs[idx].cxt_len == specs[idx + 1].cxt_len
        ):
            roles[idx] = (True, False)
            roles[idx + 1] = (False, True)
    return roles


def iter_operation_rows(rows: Iterable[Dict[str, Any]]) -> Iterable[Tuple[Dict[str, Any], Dict[str, Any]]]:
    context: Dict[str, Any] = {}
    last_model = None
    last_layer = None
    context_names = ["BS", "CXT_LEN", "r", "in_feature", "out_feature", "svd_rank"]

    for row in rows:
        if _nonempty(row.get("model")):
            last_model = row["model"]
        if _nonempty(row.get("layer")):
            last_layer = row["layer"]
        for name in context_names:
            if _nonempty(row.get(name)):
                context[name] = row[name]

        linear = row.get("linear")
        if not _nonempty(linear):
            continue
        linear_name = str(linear).strip()
        if linear_name in {"\\", "/"}:
            continue
        if linear_name.lower() not in {"w*x", "wx", "w*dx", "wdx", "svd_out", "svd_x", "svd_dx"}:
            continue

        op_context = dict(context)
        op_context["model"] = last_model
        op_context["layer"] = last_layer
        op_context["linear"] = linear_name
        for name in ("w_prec", "i_prec", "o_prec"):
            op_context[name] = row.get(name)
        yield row, op_context


@dataclass(frozen=True)
class FusedStage:
    spec: Optional[GemmSpec]
    name: str
    role: str
    remove_dram_components: Tuple[str, ...] = ()


V2_PACK_STAGE_NAMES = ("fused_v_x_rank_tmp", "fused_v_plus_dv_x_rank_tmp")
V2_PACK_GROUP = "v2_x_stream"


def _normalized_group_key(ctx: Dict[str, Any]) -> Tuple[str, str]:
    return (str(ctx["model"]).strip(), str(ctx["layer"]).strip())


def collect_operation_groups(
    rows: Iterable[Dict[str, Any]]
) -> List[Tuple[Tuple[str, str], Dict[str, Tuple[Dict[str, Any], Dict[str, Any]]]]]:
    groups: Dict[Tuple[str, str], Dict[str, Tuple[Dict[str, Any], Dict[str, Any]]]] = {}
    order: List[Tuple[str, str]] = []
    for source_row, ctx in iter_operation_rows(rows):
        key = _normalized_group_key(ctx)
        if key not in groups:
            groups[key] = {}
            order.append(key)
        groups[key][str(ctx["linear"]).strip().lower()] = (source_row, ctx)
    return [(key, groups[key]) for key in order]


def _main_spec_from_ctx(
    name: str, ctx: Dict[str, Any], in_features: int, out_features: int
) -> GemmSpec:
    return GemmSpec(
        name=name,
        batch_size=_to_int(ctx["BS"], "BS"),
        cxt_len=_to_int(ctx["CXT_LEN"], "CXT_LEN"),
        in_features=in_features,
        out_features=out_features,
        w_prec=_to_float(ctx["w_prec"], "w_prec"),
        i_prec=_to_float(ctx["i_prec"], "i_prec"),
        output_prec=_to_float(ctx["o_prec"], "o_prec"),
    )


def _factor_spec_from_ctx(
    name: str, ctx: Dict[str, Any], rows: int, cols: int
) -> GemmSpec:
    return GemmSpec(
        name=name,
        batch_size=1,
        cxt_len=rows,
        in_features=_to_int(ctx["r"], "r"),
        out_features=cols,
        w_prec=_to_float(ctx["w_prec"], "w_prec"),
        i_prec=_to_float(ctx["i_prec"], "i_prec"),
        output_prec=_to_float(ctx["o_prec"], "o_prec"),
    )


def build_dual_output_fused_stages(
    op_group: Dict[str, Tuple[Dict[str, Any], Dict[str, Any]]]
) -> Tuple[Dict[str, Any], List[FusedStage]]:
    required = ["w*x", "w*dx", "svd_out", "svd_x", "svd_dx"]
    missing = [name for name in required if name not in op_group]
    if missing:
        model = next(iter(op_group.values()))[1].get("model", "<unknown>") if op_group else "<unknown>"
        layer = next(iter(op_group.values()))[1].get("layer", "<unknown>") if op_group else "<unknown>"
        raise ValueError(f"Missing operations for fused group {model}/{layer}: {missing}")

    _, wx = op_group["w*x"]
    _, wdx = op_group["w*dx"]
    _, svd_out = op_group["svd_out"]
    _, svd_x = op_group["svd_x"]
    _, svd_dx = op_group["svd_dx"]

    in_feature = _to_int(wx["in_feature"], "in_feature")
    out_feature = _to_int(wx["out_feature"], "out_feature")
    svd_rank = _to_int(wx["svd_rank"], "svd_rank")

    stages = [
        FusedStage(
            _factor_spec_from_ctx("fused_make_dv_once", svd_x, in_feature, svd_rank),
            "fused_make_dv_once",
            "Optional factorized-dV provider; prepares V+dV tiles before Phase A.",
            (),
        ),
        FusedStage(
            _main_spec_from_ctx("fused_v_x_rank_tmp", svd_out, in_feature, svd_rank),
            "fused_v_x_rank_tmp",
            "Phase A: load x_tile and V_tile, then accumulate T_uv = x_tile @ V_tile; rank tmp stays on chip.",
            ("output",),
        ),
        FusedStage(
            _main_spec_from_ctx("fused_v_plus_dv_x_rank_tmp", svd_x, in_feature, svd_rank),
            "fused_v_plus_dv_x_rank_tmp",
            "Phase A: reuse x_tile loaded for Vx, load V+dV tile, and accumulate T_vp += x_tile @ Vp_tile.",
            ("input", "output"),
        ),
        FusedStage(
            _main_spec_from_ctx("fused_v_plus_dv_dx_rank_tmp", svd_dx, in_feature, svd_rank),
            "fused_v_plus_dv_dx_rank_tmp",
            "Phase A: reuse V+dV tile, load dx_tile, and accumulate T_vp += dx_tile @ Vp_tile.",
            ("weight", "output"),
        ),
        FusedStage(
            _main_spec_from_ctx("fused_w_x_to_forward_acc", wx, in_feature, out_feature),
            "fused_w_x_to_forward_acc",
            "Phase B: load W_tile and x_tile, then accumulate output_acc += x_tile @ W_tile; output tile stays on chip.",
            ("output",),
        ),
        FusedStage(
            _main_spec_from_ctx("fused_w_dx_to_delta_acc", wdx, in_feature, out_feature),
            "fused_w_dx_to_delta_acc",
            "Phase B: reuse W_tile loaded for Wx, load dx_tile, then accumulate diff_output_acc += dx_tile @ W_tile.",
            ("weight", "output"),
        ),
        FusedStage(
            _main_spec_from_ctx("fused_u_tmp_shared_plus_minus", svd_out, svd_rank, out_feature),
            "fused_u_tmp_shared_plus_minus",
            "Phase C: load U_tile, consume T_uv on chip, and bypass svd_output to +output_acc and -diff_output_acc.",
            ("input", "output"),
        ),
        FusedStage(
            _factor_spec_from_ctx("fused_make_du_once", svd_x, out_feature, svd_rank),
            "fused_make_du_once",
            "Optional factorized-dU provider; prepares U+dU tiles before the U' consumer.",
            (),
        ),
        FusedStage(
            _main_spec_from_ctx("fused_u_plus_du_tmp_sum_delta", svd_x, svd_rank, out_feature),
            "fused_u_plus_du_tmp_sum_delta",
            "Phase C: load U+dU tile, consume T_vp on chip, and accumulate diff_output_acc += T_vp @ Up_tile.",
            ("input", "output"),
        ),
    ]

    summary_ctx = dict(wx)
    summary_ctx["linear"] = "dual_output_fused"
    return summary_ctx, stages


def final_output_write_metrics(
    ctx: Dict[str, Any],
    effective_bw_bits_per_cycle: float,
    energy_pj_per_bit: float,
    output_count: int = 2,
) -> Dict[str, float]:
    output_bytes = (
        _to_int(ctx["BS"], "BS")
        * _to_int(ctx["CXT_LEN"], "CXT_LEN")
        * _to_int(ctx["out_feature"], "out_feature")
        * _to_float(ctx["o_prec"], "o_prec")
        / 8.0
        * output_count
    )
    dram_cycles = dram_cycles_for_bytes(output_bytes, effective_bw_bits_per_cycle)
    dram_energy = dram_energy_uJ_for_bytes(output_bytes, energy_pj_per_bit)
    return {
        "compute_latency(cycles)": 0.0,
        "dram_latency": dram_cycles,
        "total latency": dram_cycles,
        "compute energy(uJ)": 0.0,
        "sram rd/wr energy": 0.0,
        "dram energy": dram_energy,
        "on-chip energy": 0.0,
        "total energy": dram_energy,
    }


def low_rank_v2_pack_credit(
    stage_metrics: Dict[str, Dict[str, float]],
    svd_rank: int,
    pe_array_h: int,
) -> Optional[Dict[str, float]]:
    """
    Conservative V-side packing:
      pass 1 broadcasts one x stream and computes x@V plus x@Vp on disjoint
      output-channel partitions. dx@Vp is intentionally left as a separate pass.

    The two packed stages still load both V/Vp weight tensors and keep the same
    arithmetic work. The credit only removes the extra sequential pass latency.
    """

    if 2 * int(svd_rank) > int(pe_array_h):
        return None
    if any(name not in stage_metrics for name in V2_PACK_STAGE_NAMES):
        return None

    first = stage_metrics[V2_PACK_STAGE_NAMES[0]]
    second = stage_metrics[V2_PACK_STAGE_NAMES[1]]

    separate_compute = (
        first["compute_latency(cycles)"] + second["compute_latency(cycles)"]
    )
    packed_compute = max(
        first["compute_latency(cycles)"], second["compute_latency(cycles)"]
    )

    separate_total = first["total latency"] + second["total latency"]
    packed_dram = first["dram_latency"] + second["dram_latency"]
    packed_total = max(packed_compute, packed_dram)

    compute_credit = max(0.0, separate_compute - packed_compute)
    total_credit = max(0.0, separate_total - packed_total)
    if compute_credit == 0.0 and total_credit == 0.0:
        return None

    credit = empty_metrics()
    credit["compute_latency(cycles)"] = -compute_credit
    credit["total latency"] = -total_credit
    return credit


def write_csv(path: Path, rows: Sequence[Dict[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def fill_xlsx(workbook: Any, sheet: Any, headers: Sequence[str], row_to_metrics: Dict[int, Dict[str, float]], output_path: Path) -> None:
    header_to_col = {name: idx + 1 for idx, name in enumerate(headers)}
    for result_col in RESULT_COLUMNS:
        if result_col not in header_to_col:
            raise ValueError(f"Missing result column in workbook: {result_col}")
    for row_idx, metrics in row_to_metrics.items():
        for col_name in RESULT_COLUMNS:
            sheet.cell(row_idx, header_to_col[col_name]).value = metrics[col_name]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def default_output_prefix(input_path: Path) -> Path:
    return input_path.with_name(input_path.stem + "_qdiff_svd_hardware")


def run_dual_output_fused_schedule(
    args: argparse.Namespace,
    rows: Sequence[Dict[str, Any]],
    summary_csv: Path,
    detail_csv: Path,
) -> None:
    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []
    effective_bw = effective_dram_bw_bits_per_cycle(args)

    for _, op_group in collect_operation_groups(rows):
        summary_ctx, stages = build_dual_output_fused_stages(op_group)
        source_rows = [int(item[0]["__rownum__"]) for item in op_group.values()]
        excel_row = min(source_rows)
        total = empty_metrics()
        stage_metrics_by_name: Dict[str, Dict[str, float]] = {}

        for stage_idx, stage in enumerate(stages, start=1):
            if stage.spec is None:
                continue
            result = simulate_gemm(stage.spec, args)
            metrics = metrics_from_result(result)
            if stage.remove_dram_components:
                metrics = apply_dram_component_adjustment(
                    result,
                    metrics,
                    effective_bw,
                    args.dram_energy_pj_per_bit,
                    stage.remove_dram_components,
                )
            add_metrics(total, metrics)
            stage_metrics_by_name[stage.name] = metrics

            detail = {
                "excel_row": excel_row,
                "model": summary_ctx["model"],
                "layer": summary_ctx["layer"],
                "linear": "dual_output_fused",
                "schedule": args.schedule,
                "low_rank_pack": args.low_rank_pack,
                "pack_group": (
                    V2_PACK_GROUP
                    if args.low_rank_pack == "v2" and stage.name in V2_PACK_STAGE_NAMES
                    else ""
                ),
                "gemm_index": stage_idx,
                "gemm_name": stage.name,
                "stage_role": stage.role,
                "removed_dram_components": metrics.get(
                    "_removed_components", "+".join(stage.remove_dram_components)
                ),
                "BS": stage.spec.batch_size,
                "CXT_LEN": stage.spec.cxt_len,
                "IN_FEATURES": stage.spec.in_features,
                "OUT_FEATURES": stage.spec.out_features,
                "w_prec": stage.spec.w_prec,
                "i_prec": stage.spec.i_prec,
                "o_prec": stage.spec.output_prec,
                "gemm_M": result["gemm_shape"]["m"],
                "gemm_N": result["gemm_shape"]["n"],
                "gemm_K": result["gemm_shape"]["k"],
                "uv_tmp_location": "sram",
                "tmp_output_dram_bypass": int("output" in stage.remove_dram_components),
                "tmp_input_dram_bypass": int("input" in stage.remove_dram_components),
                "dram_latency_removed": metrics.get("_dram_latency_removed", 0.0),
                "total_latency_removed": metrics.get("_total_latency_removed", 0.0),
                "dram_energy_removed(uJ)": metrics.get("_dram_energy_removed", 0.0),
            }
            detail.update(metrics)
            for internal_key in (
                "_dram_latency_removed",
                "_total_latency_removed",
                "_dram_energy_removed",
                "_removed_components",
            ):
                detail.pop(internal_key, None)
            detail_rows.append(detail)

        next_detail_index = len(stages) + 1
        if args.low_rank_pack == "v2":
            pack_credit = low_rank_v2_pack_credit(
                stage_metrics_by_name,
                _to_int(summary_ctx["svd_rank"], "svd_rank"),
                args.pe_array_h,
            )
            if pack_credit is not None:
                add_metrics(total, pack_credit)
                detail = {
                    "excel_row": excel_row,
                    "model": summary_ctx["model"],
                    "layer": summary_ctx["layer"],
                    "linear": "dual_output_fused",
                    "schedule": args.schedule,
                    "low_rank_pack": args.low_rank_pack,
                    "pack_group": V2_PACK_GROUP,
                    "gemm_index": next_detail_index,
                    "gemm_name": "low_rank_v2_x_stream_pack_credit",
                    "stage_role": (
                        "Conservative V-side packing credit: one x stream computes "
                        "x@V and x@Vp on disjoint output-channel partitions; dx@Vp "
                        "remains a separate pass."
                    ),
                    "removed_dram_components": "",
                    "BS": summary_ctx["BS"],
                    "CXT_LEN": summary_ctx["CXT_LEN"],
                    "IN_FEATURES": summary_ctx["in_feature"],
                    "OUT_FEATURES": summary_ctx["svd_rank"],
                    "w_prec": "",
                    "i_prec": "",
                    "o_prec": summary_ctx["o_prec"],
                    "gemm_M": "",
                    "gemm_N": "",
                    "gemm_K": "",
                    "uv_tmp_location": "sram",
                    "tmp_output_dram_bypass": 0,
                    "tmp_input_dram_bypass": 0,
                    "dram_latency_removed": 0.0,
                    "total_latency_removed": -pack_credit["total latency"],
                    "dram_energy_removed(uJ)": 0.0,
                }
                detail.update(pack_credit)
                detail_rows.append(detail)
                next_detail_index += 1

        final_metrics = final_output_write_metrics(
            summary_ctx,
            effective_bw,
            args.dram_energy_pj_per_bit,
            output_count=2,
        )
        add_metrics(total, final_metrics)
        detail = {
            "excel_row": excel_row,
            "model": summary_ctx["model"],
            "layer": summary_ctx["layer"],
            "linear": "dual_output_fused",
            "schedule": args.schedule,
            "low_rank_pack": args.low_rank_pack,
            "pack_group": "",
            "gemm_index": next_detail_index,
            "gemm_name": "final_forward_delta_output_write",
            "stage_role": "Write forward_out and delta_out to DRAM once each after on-chip accumulation.",
            "removed_dram_components": "",
            "BS": summary_ctx["BS"],
            "CXT_LEN": summary_ctx["CXT_LEN"],
            "IN_FEATURES": "",
            "OUT_FEATURES": summary_ctx["out_feature"],
            "w_prec": "",
            "i_prec": "",
            "o_prec": summary_ctx["o_prec"],
            "gemm_M": "",
            "gemm_N": "",
            "gemm_K": "",
            "uv_tmp_location": "sram",
            "tmp_output_dram_bypass": 0,
            "tmp_input_dram_bypass": 0,
            "dram_latency_removed": 0.0,
            "total_latency_removed": 0.0,
            "dram_energy_removed(uJ)": 0.0,
        }
        detail.update(final_metrics)
        detail_rows.append(detail)

        summary = {
            "excel_row": excel_row,
            "model": summary_ctx["model"],
            "layer": summary_ctx["layer"],
            "linear": "dual_output_fused",
            "schedule": args.schedule,
            "low_rank_pack": args.low_rank_pack,
            "BS": summary_ctx["BS"],
            "CXT_LEN": summary_ctx["CXT_LEN"],
            "r": summary_ctx["r"],
            "in_feature": summary_ctx["in_feature"],
            "out_feature": summary_ctx["out_feature"],
            "svd_rank": summary_ctx["svd_rank"],
            "w_prec": "",
            "i_prec": "",
            "o_prec": summary_ctx["o_prec"],
            "gemm_count": len(stages),
            "uv_tmp_location": "sram",
            "dram_profile": args.dram_profile,
            "dram_effective_bw_bits_per_cycle": effective_bw,
        }
        summary.update(total)
        summary_rows.append(summary)

        print(
            f'{summary_ctx["model"]} / {summary_ctx["layer"]}: dual-output fused -> '
            f'{len(stages)} GEMM stages + final writes, total latency={total["total latency"]:.0f}, '
            f'total energy={total["total energy"]:.6f} uJ'
        )

    summary_fields = [
        "excel_row",
        "model",
        "layer",
        "linear",
        "schedule",
        "low_rank_pack",
        "BS",
        "CXT_LEN",
        "r",
        "in_feature",
        "out_feature",
        "svd_rank",
        "w_prec",
        "i_prec",
        "o_prec",
        "gemm_count",
        "uv_tmp_location",
        "dram_profile",
        "dram_effective_bw_bits_per_cycle",
        *RESULT_COLUMNS,
    ]
    detail_fields = [
        "excel_row",
        "model",
        "layer",
        "linear",
        "schedule",
        "low_rank_pack",
        "pack_group",
        "gemm_index",
        "gemm_name",
        "stage_role",
        "removed_dram_components",
        "BS",
        "CXT_LEN",
        "IN_FEATURES",
        "OUT_FEATURES",
        "w_prec",
        "i_prec",
        "o_prec",
        "gemm_M",
        "gemm_N",
        "gemm_K",
        "uv_tmp_location",
        "tmp_output_dram_bypass",
        "tmp_input_dram_bypass",
        "dram_latency_removed",
        "total_latency_removed",
        "dram_energy_removed(uJ)",
        *RESULT_COLUMNS,
    ]
    write_csv(summary_csv, summary_rows, summary_fields)
    write_csv(detail_csv, detail_rows, detail_fields)
    print(f"Wrote summary CSV: {summary_csv}")
    print(f"Wrote detail CSV:  {detail_csv}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Input .xlsx/.xlsm/.csv table.")
    parser.add_argument("--sheet", default=None, help="Worksheet name for xlsx input. Defaults to active sheet.")
    parser.add_argument("--output-prefix", type=Path, default=None, help="Output prefix without suffix.")
    parser.add_argument("--output-xlsx", type=Path, default=None, help="Optional filled xlsx output path.")
    parser.add_argument("--no-xlsx", action="store_true", help="Do not write an xlsx copy even when input is xlsx.")
    parser.add_argument("--kv-prec", type=float, default=8)
    parser.add_argument("--pe-dp-size", type=int, default=1)
    parser.add_argument("--pe-energy", type=float, default=0.56)
    parser.add_argument("--pe-area", type=float, default=1507.7)
    parser.add_argument("--pe-array-h", type=int, default=32)
    parser.add_argument("--pe-array-w", type=int, default=32)
    parser.add_argument("--base-activation-prec", type=float, default=16)
    parser.add_argument("--base-weight-prec", type=float, default=16)
    parser.add_argument("--is-bit-serial", action="store_true")
    parser.add_argument(
        "--schedule",
        choices=("separate", "dual-output-fused"),
        default="separate",
        help=(
            "separate preserves the original per-row GEMM decomposition; "
            "dual-output-fused models one fused forward/delta schedule per model/layer."
        ),
    )
    parser.add_argument(
        "--low-rank-pack",
        choices=("none", "v2"),
        default="none",
        help=(
            "Optional low-rank packing model for --schedule dual-output-fused. "
            "v2 conservatively packs x@V and x@Vp into one x-stream pass; dx@Vp "
            "still runs separately."
        ),
    )
    parser.add_argument(
        "--backend",
        choices=("script", "direct"),
        default="script",
        help="script calls single_linear_sim_bitFusion.py for every GEMM; direct uses the Python class API.",
    )
    parser.add_argument(
        "--sim-script",
        type=Path,
        default=Path(__file__).with_name("single_linear_sim_bitFusion.py"),
        help="Path to single_linear_sim_bitFusion.py when --backend script is used.",
    )
    parser.add_argument(
        "--uv-tmp-location",
        choices=("dram", "sram"),
        default="dram",
        help="Use sram to model Vx->Utmp with tmp kept on chip instead of written/read through DRAM.",
    )
    parser.add_argument(
        "--dram-profile",
        choices=("legacy", *DRAM_PROFILE_DATA_RATE_MTPS.keys()),
        default="legacy",
    )
    parser.add_argument("--dram-data-rate-mtps", type=float, default=None)
    parser.add_argument("--dram-bus-width-bits", type=float, default=64.0)
    parser.add_argument("--dram-channels", type=int, default=1)
    parser.add_argument("--accelerator-freq-mhz", type=float, default=1000.0)
    parser.add_argument("--dram-energy-pj-per-bit", type=float, default=18.75)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve()
    output_prefix = (args.output_prefix or default_output_prefix(input_path)).resolve()
    summary_csv = output_prefix.with_name(output_prefix.name + "_summary.csv")
    detail_csv = output_prefix.with_name(output_prefix.name + "_detail.csv")

    headers, rows, workbook, sheet = read_table(input_path, args.sheet)
    missing = [name for name in PARAM_COLUMNS if name not in headers]
    if missing:
        raise SystemExit(f"Input table is missing required columns: {missing}")

    if args.schedule == "dual-output-fused":
        run_dual_output_fused_schedule(args, rows, summary_csv, detail_csv)
        if workbook is not None and not args.no_xlsx:
            print(
                "Skipping filled XLSX for --schedule dual-output-fused because the output "
                "has one fused row per model/layer instead of one row per original operation."
            )
        return

    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []
    row_to_metrics: Dict[int, Dict[str, float]] = {}

    for source_row, ctx in iter_operation_rows(rows):
        specs = build_gemm_specs(ctx["linear"], ctx)
        total = empty_metrics()
        tmp_sram_roles = uv_tmp_sram_roles(specs) if args.uv_tmp_location == "sram" else {}

        for gemm_idx, spec in enumerate(specs, start=1):
            result = simulate_gemm(spec, args)
            metrics = metrics_from_result(result)
            remove_output_tmp, remove_input_tmp = tmp_sram_roles.get(gemm_idx - 1, (False, False))
            if remove_output_tmp or remove_input_tmp:
                metrics = apply_uv_tmp_sram_adjustment(
                    spec,
                    result,
                    metrics,
                    effective_bw_bits_per_cycle=effective_dram_bw_bits_per_cycle(args),
                    energy_pj_per_bit=args.dram_energy_pj_per_bit,
                    remove_output_tmp=remove_output_tmp,
                    remove_input_tmp=remove_input_tmp,
                )
            add_metrics(total, metrics)
            detail = {
                "excel_row": source_row["__rownum__"],
                "model": ctx["model"],
                "layer": ctx["layer"],
                "linear": ctx["linear"],
                "gemm_index": gemm_idx,
                "gemm_name": spec.name,
                "BS": spec.batch_size,
                "CXT_LEN": spec.cxt_len,
                "IN_FEATURES": spec.in_features,
                "OUT_FEATURES": spec.out_features,
                "w_prec": spec.w_prec,
                "i_prec": spec.i_prec,
                "o_prec": spec.output_prec,
                "gemm_M": result["gemm_shape"]["m"],
                "gemm_N": result["gemm_shape"]["n"],
                "gemm_K": result["gemm_shape"]["k"],
                "uv_tmp_location": args.uv_tmp_location,
                "tmp_output_dram_bypass": int(remove_output_tmp),
                "tmp_input_dram_bypass": int(remove_input_tmp),
                "dram_latency_removed": metrics.get("_dram_latency_removed", 0.0),
                "total_latency_removed": metrics.get("_total_latency_removed", 0.0),
                "dram_energy_removed(uJ)": metrics.get("_dram_energy_removed", 0.0),
            }
            detail.update(metrics)
            for internal_key in ("_dram_latency_removed", "_total_latency_removed", "_dram_energy_removed"):
                detail.pop(internal_key, None)
            detail_rows.append(detail)

        summary = {
            "excel_row": source_row["__rownum__"],
            "model": ctx["model"],
            "layer": ctx["layer"],
            "linear": ctx["linear"],
            "BS": ctx["BS"],
            "CXT_LEN": ctx["CXT_LEN"],
            "r": ctx["r"],
            "in_feature": ctx["in_feature"],
            "out_feature": ctx["out_feature"],
            "svd_rank": ctx["svd_rank"],
            "w_prec": ctx["w_prec"],
            "i_prec": ctx["i_prec"],
            "o_prec": ctx["o_prec"],
            "gemm_count": len(specs),
            "uv_tmp_location": args.uv_tmp_location,
            "dram_profile": args.dram_profile,
            "dram_effective_bw_bits_per_cycle": effective_dram_bw_bits_per_cycle(args),
        }
        summary.update(total)
        summary_rows.append(summary)
        row_to_metrics[int(source_row["__rownum__"])] = total

        print(
            f'row {source_row["__rownum__"]}: {ctx["model"]} / {ctx["layer"]} / '
            f'{ctx["linear"]} -> {len(specs)} GEMMs, total latency={total["total latency"]:.0f}, '
            f'total energy={total["total energy"]:.6f} uJ'
        )

    summary_fields = [
        "excel_row",
        "model",
        "layer",
        "linear",
        "BS",
        "CXT_LEN",
        "r",
        "in_feature",
        "out_feature",
        "svd_rank",
        "w_prec",
        "i_prec",
        "o_prec",
        "gemm_count",
        "uv_tmp_location",
        "dram_profile",
        "dram_effective_bw_bits_per_cycle",
        *RESULT_COLUMNS,
    ]
    detail_fields = [
        "excel_row",
        "model",
        "layer",
        "linear",
        "gemm_index",
        "gemm_name",
        "BS",
        "CXT_LEN",
        "IN_FEATURES",
        "OUT_FEATURES",
        "w_prec",
        "i_prec",
        "o_prec",
        "gemm_M",
        "gemm_N",
        "gemm_K",
        "uv_tmp_location",
        "tmp_output_dram_bypass",
        "tmp_input_dram_bypass",
        "dram_latency_removed",
        "total_latency_removed",
        "dram_energy_removed(uJ)",
        *RESULT_COLUMNS,
    ]
    write_csv(summary_csv, summary_rows, summary_fields)
    write_csv(detail_csv, detail_rows, detail_fields)
    print(f"Wrote summary CSV: {summary_csv}")
    print(f"Wrote detail CSV:  {detail_csv}")

    if workbook is not None and not args.no_xlsx:
        xlsx_path = args.output_xlsx or output_prefix.with_name(output_prefix.name + "_filled.xlsx")
        fill_xlsx(workbook, sheet, headers, row_to_metrics, xlsx_path.resolve())
        print(f"Wrote filled XLSX: {xlsx_path.resolve()}")


if __name__ == "__main__":
    main()

